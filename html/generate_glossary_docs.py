import json
import re
import os
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Configuration
JSON_FILE = 'Abkürzungen und Bedeutungen.json'
PDF_FILE = 'Abkürzungen_und_Bedeutungen.pdf'
EXCEL_FILE = 'Abkürzungen_und_Bedeutungen.xlsx'

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
JSON_PATH = os.path.join(SCRIPT_DIR, JSON_FILE)
PDF_PATH = os.path.join(SCRIPT_DIR, PDF_FILE)
EXCEL_PATH = os.path.join(SCRIPT_DIR, EXCEL_FILE)

# Paths for scanning exercises
SEARCH_DIRS = [
    os.path.join(SCRIPT_DIR, r"../docs/training1/Ventilsteuerung/4diacIDE-workspace/test_B/Uebungen"),
    os.path.join(SCRIPT_DIR, r"../docs/training1/Ventilsteuerung/4diacIDE-workspace/test_AX/Uebungen")
]

def clean_html(raw_html):
    """Remove HTML tags from a string."""
    if not raw_html:
        return ""
    cleanr = re.compile('<.*?>')
    cleantext = re.sub(cleanr, '', str(raw_html))
    return cleantext.strip()

def load_data():
    if not os.path.exists(JSON_PATH):
        print(f"Error: {JSON_PATH} not found.")
        return None
    with open(JSON_PATH, 'r', encoding='utf-8') as f:
        return json.load(f)

def save_data(data):
    with open(JSON_PATH, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=3, ensure_ascii=False)
    print(f"Saved updated JSON to {JSON_PATH}")

def scan_for_all_exercises():
    """Scans for all FB usage in exercises. Returns a dict {FB_NAME: [EX1, EX2, ...]}"""
    # We collect all terms from the JSON to know what to look for
    data = load_data()
    if not data: return {}
    
    target_fbs = []
    for cat in data.get('categories', []):
        for item in cat.get('data', []):
            term = item.get('term', '')
            # Handle terms like "E_T_FF / Flip-Flop"
            clean_term = term.split('/')[0].strip()
            if clean_term and clean_term not in target_fbs:
                target_fbs.append(clean_term)

    fb_usage = {}
    for directory in SEARCH_DIRS:
        if not os.path.exists(directory):
            print(f"Directory not found for scanning: {directory}")
            continue

        for filename in os.listdir(directory):
            if filename.endswith(".SUB"):
                filepath = os.path.join(directory, filename)
                exercise_name = os.path.splitext(filename)[0]
                
                try:
                    with open(filepath, 'r', encoding='utf-8') as f:
                        content = f.read()
                        for fb in target_fbs:
                            pattern = fr'Type="[^"]*([:_]){re.escape(fb)}"'
                            pattern_exact = fr'Type="{re.escape(fb)}"'
                            if re.search(pattern, content) or re.search(pattern_exact, content):
                                if fb not in fb_usage:
                                    fb_usage[fb] = []
                                if exercise_name not in fb_usage[fb]:
                                    fb_usage[fb].append(exercise_name)
                except Exception as e:
                    print(f"Error reading {filepath}: {e}")
    return fb_usage

def update_json_with_exercises():
    print("Scanning exercises to update JSON...")
    usage = scan_for_all_exercises()
    data = load_data()
    if not data: return

    base_url = "https://meisterschulen-am-ostbahnhof-munchen-docs.readthedocs.io/projects/visual-programming-languages-docs/de/latest/training1/Ventilsteuerung/4diacIDE-workspace/"

    for cat in data.get('categories', []):
        for item in cat.get('data', []):
            term = item.get('term', '')
            clean_term = term.split('/')[0].strip()
            
            if clean_term in usage:
                exercises = usage[clean_term]
                links = []
                for ex in sorted(exercises):
                    # Determine subfolder based on exercise name or search dir
                    # For simplicity, we check if it exists in test_B or test_AX
                    subfolder = "test_B"
                    if "_AX" in ex:
                        subfolder = "test_AX"
                    
                    url = f"{base_url}{subfolder}/Uebungen_doc/{ex}.html"
                    links.append(f'<a href="{url}" target="_blank">{ex}</a>')
                
                item['ex'] = ", ".join(links)

    save_data(data)

def generate_excel(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Glossar"

    # Headers
    headers = ["Kategorie", "Nr", "Begriff", "Bedeutung", "Titel", "Typ", "Beispiele", "Links (Intern)"]
    ws.append(headers)

    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="9A5130", end_color="9A5130", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for category in data.get('categories', []):
        cat_title = category.get('title', 'Unkown')
        for item in category.get('data', []):
            row = [
                cat_title,
                item.get('nr', ''),
                item.get('term', ''),
                item.get('mean', ''),
                item.get('title', ''),
                item.get('type', ''),
                item.get('ex', ''),
                clean_html(item.get('link_int', ''))
            ]
            ws.append(row)

    # Auto-adjust column widths (rough estimate)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        if adjusted_width > 50: adjusted_width = 50 # Cap width
        ws.column_dimensions[column].width = adjusted_width

    print(f"Saving Excel to {EXCEL_PATH}...")
    wb.save(EXCEL_PATH)

def generate_pdf(data):
    doc = SimpleDocTemplate(PDF_PATH, pagesize=landscape(A4), topMargin=1*cm, bottomMargin=1*cm)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    title_style.textColor = colors.HexColor("#733C24")
    
    h2_style = styles['Heading2']
    h2_style.textColor = colors.HexColor("#9A5130")
    h2_style.spaceBefore = 12
    h2_style.spaceAfter = 6

    # Normal text for table cells
    cell_style = ParagraphStyle(
        'CellStyle',
        parent=styles['Normal'],
        fontSize=9,
        leading=11
    )
    
    # Title
    elements.append(Paragraph("Abkürzungen und Bedeutungen", title_style))
    elements.append(Spacer(1, 0.5*cm))

    # Table columns: Term, Meaning, Title, Examples
    # We omit some less important ones or combine them to fit
    
    for category in data.get('categories', []):
        cat_title = category.get('title', 'Unkown')
        elements.append(Paragraph(cat_title, h2_style))

        table_data = [["Begriff", "Bedeutung", "Titel / Beschreibung", "Typ / Beispiele"]]
        
        for item in category.get('data', []):
            term = Paragraph(item.get('term', ''), cell_style)
            mean = Paragraph(item.get('mean', ''), cell_style)
            
            # Combine Title and basic description if needed, currently using 'title' field
            title_text = item.get('title', '')
            title_p = Paragraph(title_text, cell_style)
            
            # Combine Type and Examples
            type_str = item.get('type', '')
            ex_str = item.get('ex', '')
            
            # Limit examples in PDF to prevent extreme row height
            ex_list = ex_str.split(', ')
            if len(ex_list) > 5:
                ex_str = ", ".join(ex_list[:5]) + " ..."
            
            combined_ex = []
            if type_str: combined_ex.append(f"<b>Typ:</b> {type_str}")
            if ex_str: combined_ex.append(f"<b>Bsp:</b> {ex_str}")
            
            ex_p = Paragraph("<br/>".join(combined_ex), cell_style)

            table_data.append([term, mean, title_p, ex_p])

        if len(table_data) > 1:
            # Create Table
            # Widths: Term (15%), Mean (30%), Title (25%), Ex (30%)
            page_width = landscape(A4)[0] - 2*cm
            col_widths = [page_width * 0.15, page_width * 0.30, page_width * 0.25, page_width * 0.30]
            
            t = Table(table_data, colWidths=col_widths, repeatRows=1, splitByRow=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#f0f0f0")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ]))
            elements.append(t)
            elements.append(Spacer(1, 0.5*cm))
        else:
            elements.append(Paragraph("(Keine Einträge)", cell_style))

    print(f"Saving PDF to {PDF_PATH}...")
    doc.build(elements)

def main():
    print("Starting generation...")
    update_json_with_exercises()
    data = load_data()
    if data:
        generate_excel(data)
        generate_pdf(data)
        print("Generation complete.")
    else:
        print("Failed to load data.")

if __name__ == "__main__":
    main()
